import streamlit as st
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
import streamlit.components.v1 as components

# Configuración de la página
st.set_page_config(
    page_title="DaTASA - Control de Proyectos",
    page_icon="🚢",
    layout="wide",
    initial_sidebar_state="expanded"
)

st.title(":ship: Control de Proyectos")

# Carga de datos desde Excel
BD = './BD.xlsx'
try:
    TEMPORADAS = load_workbook(BD, read_only=True).sheetnames
except Exception as e:
    st.error(f"No se pudo cargar el archivo '{BD}'. Error: {e}")
    st.stop()

# Sidebar - Selección de temporada y proyecto
st.sidebar.header('Proyectos :anchor:')
selector_temporada = st.sidebar.selectbox("Seleccione Tipo de Buque:", TEMPORADAS, index=0)
df_proyecto = pd.read_excel(BD, sheet_name=selector_temporada)
selector_proyecto = st.sidebar.multiselect("Seleccione proyectos:", df_proyecto['Proyecto'].drop_duplicates())

# Columnas de interés
columnas_superiores = ["Manga", "Eslora", "Puntal", "Matricula"]
columnas_extra = ["Temporada Gestion de Astillero", "Costo Total (PEN)", "Identificador"]

# Mapeo de tipo de buque a archivo correspondiente
archivo_proyecto_map = {
    "REMOLCADOR": './REMOLCADOR.xlsx',
    "CHATA": './CHATA.xlsx',
    "EMBARCACION PESQUERA": './EMBARCACIÓN PESQUERA.xlsx',
    "PANGA": './PANGA.xlsx'
}

# Determinar archivo correspondiente
temporada_key = selector_temporada.upper().strip()
archivo_seleccionado = None
for tipo, ruta in archivo_proyecto_map.items():
    if tipo in temporada_key:
        archivo_seleccionado = ruta
        break

if not archivo_seleccionado:
    st.warning("❗ No se reconoce el tipo de buque para esta temporada.")
    st.stop()

# Mostrar detalles del proyecto
if selector_proyecto:
    for proyecto in selector_proyecto:
        try:
            df_historial = pd.read_excel(archivo_seleccionado, sheet_name=proyecto)

            # Mostrar en columnas: Modelo 3D y Detalles
            col1, col2 = st.columns([1.2, 1.8])

            with col1:
                st.subheader(f"📦 Modelo 3D - {proyecto}")
                poly_url = "https://poly.cam/capture/331664f9-ff14-4235-b19e-3b274191c49c"
                components.iframe(poly_url, height=400)

            with col2:
                columnas_1 = [col for col in columnas_superiores if col in df_historial.columns]
                if columnas_1:
                    info_general = df_historial[columnas_1].iloc[0]
                    st.subheader("📋 Información Técnica")
                    for col in columnas_1:
                        st.markdown(f"**{col}:** {info_general[col]}")

                columnas_2 = [col for col in columnas_extra if col in df_historial.columns]
                if columnas_2:
                    info_extra = df_historial[columnas_2].iloc[0:3].copy()
                    if "Costo Total (PEN)" in info_extra.columns:
                        info_extra["Costo Total (PEN)"] = info_extra["Costo Total (PEN)"].apply(lambda x: f"S/. {x:,.2f}")
                    st.subheader("🧾 Seguimiento del Proyecto")
                    st.dataframe(info_extra)

            # Historial técnico
            columnas_usadas = columnas_1 + columnas_2
            df_tecnico = df_historial.drop(columns=columnas_usadas, errors='ignore')
            if not df_tecnico.empty:
                st.subheader(f"📏 Historial metrado del proyecto: {proyecto}")
                st.dataframe(df_tecnico)
            else:
                st.info("No hay datos adicionales para mostrar en historial.")

        except Exception as e:
            st.error(f"❌ No se pudo cargar la hoja '{proyecto}' en {archivo_seleccionado}. Error: {e}")
else:
    st.info("Seleccione al menos un proyecto para mostrar información.")
